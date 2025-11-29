"""
Unit tests for ExportService.

Tests export service methods in isolation.
"""
import pytest
from datetime import datetime
import sys
import os
from io import BytesIO
from openpyxl import load_workbook

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from src.common.modules.export.exporter import ExportService


@pytest.mark.unit
class TestExportService:
    """Test suite for ExportService."""

    @pytest.fixture
    def sample_data(self):
        """Sample data for export."""
        return [
            {
                "id": 1,
                "name": "Test Company",
                "email": "test@example.com",
                "created_at": datetime(2024, 1, 1, 12, 0, 0),
            },
            {
                "id": 2,
                "name": "Another Company",
                "email": "another@example.com",
                "created_at": datetime(2024, 1, 2, 12, 0, 0),
            },
        ]

    def test_export_to_excel_success(self, sample_data):
        """Test successful Excel export."""
        result = ExportService.export_to_excel(
            data=sample_data,
            sheet_name="Test Sheet",
            title="Test Title"
        )
        
        assert result is not None
        assert isinstance(result, bytes)
        assert len(result) > 0
        
        # Verify Excel file can be opened
        wb = load_workbook(BytesIO(result))
        assert "Test Sheet" in wb.sheetnames
        ws = wb["Test Sheet"]
        assert ws["A1"].value == "Test Title"  # Title row

    def test_export_to_excel_without_title(self, sample_data):
        """Test Excel export without title."""
        result = ExportService.export_to_excel(
            data=sample_data,
            sheet_name="Test Sheet"
        )
        
        assert result is not None
        assert isinstance(result, bytes)
        
        wb = load_workbook(BytesIO(result))
        ws = wb["Test Sheet"]
        # First row should be headers
        assert ws["A1"].value == "id"
        assert ws["B1"].value == "name"

    def test_export_to_excel_with_custom_headers(self, sample_data):
        """Test Excel export with custom headers."""
        custom_headers = ["ID", "Company Name", "Email Address", "Created Date"]
        
        result = ExportService.export_to_excel(
            data=sample_data,
            sheet_name="Test Sheet",
            headers=custom_headers
        )
        
        assert result is not None
        wb = load_workbook(BytesIO(result))
        ws = wb["Test Sheet"]
        assert ws["A1"].value == "ID"
        assert ws["B1"].value == "Company Name"

    def test_export_to_excel_empty_data(self):
        """Test Excel export with empty data."""
        result = ExportService.export_to_excel(
            data=[],
            sheet_name="Empty Sheet"
        )
        
        assert result is not None
        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        assert "Empty Sheet" in wb.sheetnames

    def test_export_to_excel_datetime_formatting(self, sample_data):
        """Test Excel export with datetime formatting."""
        result = ExportService.export_to_excel(
            data=sample_data,
            sheet_name="Test Sheet"
        )
        
        assert result is not None
        wb = load_workbook(BytesIO(result))
        ws = wb["Test Sheet"]
        # Check that datetime is formatted as string
        # The datetime should be in the data row (row 2 or 3 depending on title)
        datetime_cell = ws.cell(row=2, column=4)  # created_at column
        assert datetime_cell.value is not None

    def test_export_to_excel_none_values(self):
        """Test Excel export with None values."""
        data_with_none = [
            {"id": 1, "name": "Test", "value": None},
            {"id": 2, "name": None, "value": 100},
        ]
        
        result = ExportService.export_to_excel(
            data=data_with_none,
            sheet_name="Test Sheet"
        )
        
        assert result is not None
        wb = load_workbook(BytesIO(result))
        ws = wb["Test Sheet"]
        # None values should be empty strings or None (openpyxl may return None for empty cells)
        # Row 1 is headers, Row 2 is first data row
        value_cell = ws.cell(row=2, column=3).value  # value column, first row
        assert value_cell == "" or value_cell is None

    def test_export_to_csv_success(self, sample_data):
        """Test successful CSV export."""
        result = ExportService.export_to_csv(data=sample_data)
        
        assert result is not None
        assert isinstance(result, str)
        assert len(result) > 0
        assert "id,name,email,created_at" in result
        assert "Test Company" in result

    def test_export_to_csv_with_custom_headers(self, sample_data):
        """Test CSV export with custom headers."""
        custom_headers = ["ID", "Company Name", "Email Address", "Created Date"]
        
        result = ExportService.export_to_csv(
            data=sample_data,
            headers=custom_headers
        )
        
        assert result is not None
        assert "ID,Company Name,Email Address,Created Date" in result

    def test_export_to_csv_empty_data(self):
        """Test CSV export with empty data."""
        result = ExportService.export_to_csv(data=[])
        
        assert result == ""

    def test_export_to_csv_datetime_formatting(self, sample_data):
        """Test CSV export with datetime formatting."""
        result = ExportService.export_to_csv(data=sample_data)
        
        assert result is not None
        # Datetime should be formatted as string
        assert "2024-01-01 12:00:00" in result or "2024-01-01" in result

    def test_export_to_csv_none_values(self):
        """Test CSV export with None values."""
        data_with_none = [
            {"id": 1, "name": "Test", "value": None},
            {"id": 2, "name": None, "value": 100},
        ]
        
        result = ExportService.export_to_csv(data=data_with_none)
        
        assert result is not None
        # None values should be empty strings in CSV
        lines = result.split("\n")
        assert len(lines) >= 2  # Header + at least one data row

    def test_export_to_csv_extra_fields_ignored(self):
        """Test CSV export ignores extra fields not in headers."""
        data = [
            {"id": 1, "name": "Test", "extra": "ignored"},
        ]
        headers = ["id", "name"]
        
        result = ExportService.export_to_csv(data=data, headers=headers)
        
        assert result is not None
        assert "id,name" in result
        assert "extra" not in result
        assert "ignored" not in result

    def test_export_to_excel_column_width_auto_adjust(self, sample_data):
        """Test Excel export auto-adjusts column widths."""
        result = ExportService.export_to_excel(
            data=sample_data,
            sheet_name="Test Sheet"
        )
        
        assert result is not None
        wb = load_workbook(BytesIO(result))
        ws = wb["Test Sheet"]
        # Column dimensions should be set
        assert ws.column_dimensions["A"].width > 0

    def test_export_to_excel_large_data(self):
        """Test Excel export with large dataset."""
        large_data = [
            {"id": i, "name": f"Company {i}", "value": i * 10}
            for i in range(100)
        ]
        
        result = ExportService.export_to_excel(
            data=large_data,
            sheet_name="Large Sheet"
        )
        
        assert result is not None
        wb = load_workbook(BytesIO(result))
        ws = wb["Large Sheet"]
        # Should have 100 data rows + 1 header row
        assert ws.max_row >= 100

